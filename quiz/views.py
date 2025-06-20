from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import Q
from django.contrib import messages
from .models import Quiz, Category
from quiz.models import QuizSubmission, Choice, QuizAnswer
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.utils.timezone import is_aware, make_naive, get_current_timezone
from .forms import QuizForm


# View to show all quizzes
@login_required
def all_quiz_view(request, category=None):
    quizzes = Quiz.objects.order_by('-created_at')
    categories = Category.objects.all()

    context = {"quizzes": quizzes, "categories": categories}
    return render(request, 'all-test.html', context)


# View to handle search by category or search bar
@login_required
def search_view(request, category):
    if request.GET.get('q') is not None:
        q = request.GET.get('q')
        query = Q(title__icontains=q) | Q(description__icontains=q)
        quizzes = Quiz.objects.filter(query)
    elif category != " ":
        quizzes = Quiz.objects.filter(category__name__iexact=category)
    else:
        quizzes = Quiz.objects.order_by('-created_at')

    categories = Category.objects.all()
    context = {"quizzes": quizzes, "categories": categories}
    return render(request, 'all-test.html', context)


# View to display and submit a quiz
@login_required
def quiz_view(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)

    existing_submission = QuizSubmission.objects.filter(user=request.user, quiz=quiz).first()
    if existing_submission:
        messages.info(request, "You have already attempted this quiz.")
        return redirect('quiz_result', submission_id=existing_submission.id)

    if request.method == "POST":
        score = 0
        submission = QuizSubmission.objects.create(user=request.user, quiz=quiz, score=0)

        for question in quiz.question_set.all():
            selected_id = request.POST.get(str(question.id))
            if selected_id:
                try:
                    selected_choice = Choice.objects.get(id=int(selected_id), question=question)
                except Choice.DoesNotExist:
                    selected_choice = None

                if selected_choice:
                    if selected_choice.is_correct:
                        score += 1
                    QuizAnswer.objects.create(
                        submission=submission,
                        question=question,
                        selected_choice=selected_choice
                    )

        total_questions = quiz.question_set.count()
        score = min(score, total_questions)
        submission.score = score
        submission.save()

        # Check for auto-submission and set message
        if request.POST.get("auto_submit") == "1":
            messages.warning(request, "Your quiz was auto-submitted because you switched tabs or windows.")

        return redirect('quiz_result', submission_id=submission.id)

    return render(request, 'test.html', {'quiz': quiz})

# quiz/views.py
def is_superuser_or_staff(user):
    return user.is_superuser or user.is_staff

@login_required
@user_passes_test(is_superuser_or_staff)
def add_quiz_view(request):
    if request.method == 'POST':
        form = QuizForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Quiz created and questions imported successfully!")
            return redirect('all_quiz')  # Or your quiz list page
    else:
        form = QuizForm()

    return render(request, 'add_test.html', {'form': form})

# View to show quiz result
@login_required
def quiz_result_view(request, submission_id):
    submission = get_object_or_404(QuizSubmission, id=submission_id)

    total_questions = submission.quiz.question_set.count()
    correct_answers = QuizAnswer.objects.filter(
        submission=submission,
        selected_choice__is_correct=True
    ).count()
    all_answers = QuizAnswer.objects.filter(submission=submission)

    context = {
        'submission': submission,
        'total_questions': total_questions,
        'correct_answers': correct_answers,
        'answers': all_answers,
    }

    return render(request, 'test-result.html', context)

@login_required
def quiz_report_view(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    submissions = QuizSubmission.objects.filter(quiz=quiz).order_by('-score', 'submitted_at')
   
    if not submissions.exists():
        messages.warning(request, "No submissions found for this quiz yet.")
        return redirect('all_quiz')  # Or wherever you want to send the user


    ranked_submissions = []
    last_score = None
    current_rank = 1

    for index, sub in enumerate(submissions):
        if sub.score != last_score:
            current_rank = current_rank if not ranked_submissions else ranked_submissions[-1]['rank'] + 1

        ranked_submissions.append({
            'submission': sub,
            'rank': current_rank
        })

        last_score = sub.score

    context = {
        'quiz': quiz,
        'ranked_submissions': ranked_submissions
    }
    return render(request, 'test_report.html', context)


# Export report to Excel (with correct dense ranking)
@login_required
def export_quiz_report_excel(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    total_marks = quiz.question_set.count()
    submissions = QuizSubmission.objects.filter(quiz=quiz).order_by('-score', 'submitted_at')

    ranked_data = []
    last_score = None
    current_rank = 1

    for sub in submissions:
        if sub.score != last_score:
            current_rank = current_rank if not ranked_data else ranked_data[-1][0] + 1

        submitted_at = sub.submitted_at
        if is_aware(submitted_at):
            submitted_at = make_naive(submitted_at, get_current_timezone())
          
        score_display = f"{sub.score} / {total_marks}"


        ranked_data.append([current_rank, sub.user.username, score_display, submitted_at])
        last_score = sub.score

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ['Rank', 'Username', 'Score', 'Submitted At']
    ws.append(headers)

    for row in ranked_data:
        ws.append(row)

    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{quiz.title}_Report.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response