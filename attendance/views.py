from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse,HttpResponseForbidden
from .models import Session, Attendance
from .forms import SessionForm, AttendanceForm
from django.contrib.auth.models import User
from datetime import date
import pandas as pd
from django.utils.dateparse import parse_date
from datetime import datetime
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.utils import get_column_letter



def calculate_attendance_percentage(trainee, session):
    total = Attendance.objects.filter(session=session, trainee=trainee).count()
    present = Attendance.objects.filter(session=session, trainee=trainee, present=True).count()
    return round((present / total) * 100, 2) if total > 0 else 0

def is_admin(user):
    return user.is_staff  # Or your own logic for admin

@login_required
@user_passes_test(is_admin)
def create_session(request):
    if request.method == 'POST':
        form = SessionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('attendance:session_list')
    else:
        form = SessionForm()
    return render(request, 'create_session.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def session_list(request):
    sessions = Session.objects.all()
    return render(request, 'session_list.html', {'sessions': sessions})

@login_required
@user_passes_test(is_admin)
def mark_attendance(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    trainees = session.trainees.all()
    today = date.today()

    if request.method == 'POST':
        for trainee in trainees:
            present = request.POST.get(f'present_{trainee.id}') == 'on'
            attendance, created = Attendance.objects.get_or_create(
                session=session,
                trainee=trainee,
                date=today,
                defaults={'present': present}
            )
            if not created:
                attendance.present = present
                attendance.save()
        return redirect('attendance:session_list')

    attendance_records = {a.trainee_id: a for a in Attendance.objects.filter(session=session, date=today)}
    context = {'session': session, 'trainees': trainees, 'attendance_records': attendance_records, 'date': today}
    return render(request, 'mark_attendance.html', context)


@login_required
@user_passes_test(is_admin)
def attendance_report_session(request, session_id):
    session = get_object_or_404(Session, id=session_id)

    # Base queryset: all attendance for this session
    attendance_qs = Attendance.objects.filter(session=session).order_by('-date', 'trainee__username')

    # Get filter parameters from GET
    start_date_str = request.GET.get('start')
    end_date_str = request.GET.get('end')
    trainee_id = request.GET.get('trainee')

    # If no date filters given, default to last 30 days to reduce load
    if not start_date_str and not end_date_str:
        default_start_date = date.today() - timedelta(days=30)
        attendance_qs = attendance_qs.filter(date__gte=default_start_date)

    # Apply start date filter if given
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            attendance_qs = attendance_qs.filter(date__gte=start_date)
        except ValueError:
            pass  # invalid date format, ignore filter

    # Apply end date filter if given
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            attendance_qs = attendance_qs.filter(date__lte=end_date)
        except ValueError:
            pass  # invalid date format, ignore filter

    # Apply trainee filter if given
    if trainee_id:
        attendance_qs = attendance_qs.filter(trainee__id=trainee_id)
    
    attendance = attendance_qs  

    # Calculate attendance percentages per trainee based on filtered records
   
    trainees = session.trainees.all()

    if trainee_id:
      trainees = trainees.filter(id=trainee_id)

    percentages = []
    for trainee in trainees:
        total_days = attendance_qs.filter(trainee=trainee).count()
        present_days = attendance_qs.filter(trainee=trainee, present=True).count()
        percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0
        percentages.append({'trainee': trainee, 'percentage': percentage})

    context = {
        'session': session,
        'attendance': attendance,
        'percentages': percentages,
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'selected_trainee': int(trainee_id) if trainee_id else None,
        'trainees': trainees,
    }

    return render(request, 'admin_report.html', context)

@login_required
def export_attendance_excel(request, session_id):
    session = get_object_or_404(Session, id=session_id)
   
    # Attendance records queryset
    attendance_qs = Attendance.objects.filter(session=session).select_related('trainee').order_by('date', 'trainee__username')

    # Prepare attendance data
    attendance_data = [
        {
            'Trainee': a.trainee.get_full_name() or a.trainee.username,
            'Date': a.date,
            'Present': 'Yes' if a.present else 'No'
        }
        for a in attendance_qs
    ]

    # Calculate attendance percentages per trainee
    trainees = session.trainees.all()
    percentages = []
    for trainee in trainees:
        total = attendance_qs.filter(trainee=trainee).count()
        present = attendance_qs.filter(trainee=trainee, present=True).count()
        percentage = round((present / total) * 100, 2) if total > 0 else 0
        percentages.append({'Trainee': trainee.get_full_name() or trainee.username, 'Percentage': percentage})

    # Use pandas ExcelWriter to create multiple sheets
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{session.name}_attendance.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Attendance sheet
        df_attendance = pd.DataFrame(attendance_data)
        df_attendance.to_excel(writer, index=False, sheet_name='Attendance Records')

        # Percentages sheet
        df_percentages = pd.DataFrame(percentages)
        df_percentages.to_excel(writer, index=False, sheet_name='Attendance Percentages')

    return response

@login_required
def student_attendance_report(request):
    student = request.user
 

    # Get filters
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    selected_session_id = request.GET.get('session')

    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Fix here: use 'attendances' (related_name) or 'attendance_set' if related_name not set
    attendance_records = Attendance.objects.filter(trainee=student)
    if start_date:
        attendance_records = attendance_records.filter(date__gte=start_date)
    if end_date:
        attendance_records = attendance_records.filter(date__lte=end_date)
    if selected_session_id:
        attendance_records = attendance_records.filter(session_id=selected_session_id)

    # Check if export to Excel is requested
    if request.GET.get('export') == 'excel':
        return export_attendance_to_excel(attendance_records)

    # Calculate overall attendance percentage
    total = attendance_records.count()
    present = attendance_records.filter(present=True).count()
    attendance_percentage = round((present / total) * 100, 2) if total > 0 else 0

    # Get sessions the student attended (using correct related_name 'attendances')
    sessions = Session.objects.filter(attendances__trainee=student).distinct()

    # Calculate attendance percentage per session
    individual_percentages = []
    for session in sessions:
        session_records = Attendance.objects.filter(trainee=student, session=session)
        total_in_session = session_records.count()
        present_in_session = session_records.filter(present=True).count()
        percentage = round((present_in_session / total_in_session) * 100, 2) if total_in_session > 0 else 0
        individual_percentages.append({'session': session, 'percentage': percentage})

    context = {
        'attendance_records': attendance_records.order_by('-date'),
        'attendance_percentage': attendance_percentage,
        'individual_percentages': individual_percentages,
        'session_list': sessions,
        'start_date': request.GET.get('start', ''),
        'end_date': request.GET.get('end', ''),
        'selected_session': int(selected_session_id) if selected_session_id else '',
    }

    return render(request, 'student_attendance_report.html', context)


def export_attendance_to_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Headers
    headers = ['Session', 'Date', 'Status']
    for col_num, column_title in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Data rows
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=record.session.name)
        ws.cell(row=row_num, column=2, value=str(record.date))
        ws.cell(row=row_num, column=3, value="Present" if record.present else "Absent")

    # Format columns width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'
    wb.save(response)
    return response

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseForbidden
from django.utils.dateparse import parse_date

@login_required
def student_session_report(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    user = request.user

    # Restrict access
    if not user.is_staff and user not in session.trainees.all():
        return HttpResponseForbidden("You do not have permission to view this session attendance.")

    # Get filters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    trainee_id = request.GET.get('trainee_id')

    # Base query
    attendance_records = Attendance.objects.filter(session=session)

    # Apply filters
    if start_date:
        attendance_records = attendance_records.filter(date__gte=parse_date(start_date))
    if end_date:
        attendance_records = attendance_records.filter(date__lte=parse_date(end_date))
    if trainee_id:
        attendance_records = attendance_records.filter(trainee__id=trainee_id)
        trainees = session.trainees.filter(id=trainee_id)
    else:
        trainees = session.trainees.all()

    attendance_records = attendance_records.order_by('date')

    # Calculate percentage per trainee
    percentages = []
    for trainee in trainees:
        total_days = attendance_records.filter(trainee=trainee).count()
        present_days = attendance_records.filter(trainee=trainee, present=True).count()
        percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0
        percentages.append({
            'trainee': trainee,
            'percentage': percentage,
        })

    context = {
        'session': session,
        'attendance': attendance_records,
        'percentages': percentages,
        'start_date': start_date,
        'end_date': end_date,
        'trainee_id': trainee_id,
        'trainees': session.trainees.all(),  # for dropdown filter in template
    }

    return render(request, 'student_session_report.html', context)

